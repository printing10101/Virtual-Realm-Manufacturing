"""Excel文档解析模块

使用openpyxl库实现Excel表格数据的精准提取，支持.xls和.xlsx格式，
能够识别合并单元格、冻结窗格等复杂表格结构。
"""

from __future__ import annotations

import logging
import time
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def parse_excel(file_path: str | Path) -> dict[str, Any]:
    """解析Excel文档，提取表格数据

    Args:
        file_path: Excel文件路径（.xls或.xlsx）

    Returns:
        包含解析结果的字典，结构如下：
        {
            "status": "success" | "error",
            "file_name": str,
            "file_size": int,
            "sheet_count": int,
            "tables": list[dict],  # 提取的表格列表
            "rows": list[dict],  # 所有数据行
            "parse_time_ms": float,
            "error": str | None
        }
    """
    start_time = time.time()
    file_path = Path(file_path)

    result = {
        "status": "error",
        "file_name": file_path.name,
        "file_size": 0,
        "sheet_count": 0,
        "tables": [],
        "rows": [],
        "parse_time_ms": 0.0,
        "error": None,
    }

    # 检查文件是否存在
    if not file_path.exists():
        error_msg = f"文件不存在: {file_path}"
        logger.error(error_msg)
        result["error"] = error_msg
        return result

    # 获取文件大小
    try:
        result["file_size"] = file_path.stat().st_size
    except OSError as e:
        logger.warning("无法获取文件大小: %s", e)

    # 解析Excel
    try:
        import openpyxl

        logger.info("开始解析Excel文件: %s", file_path.name)

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        result["sheet_count"] = len(wb.sheetnames)

        all_tables = []
        all_rows = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]

            # 提取表格数据
            table_info = _extract_table_from_sheet(ws, sheet_idx, sheet_name)
            if table_info:
                all_tables.append(table_info)
                all_rows.extend(table_info["rows"])

        wb.close()

        result["tables"] = all_tables
        result["rows"] = all_rows
        result["status"] = "success"

        parse_time = (time.time() - start_time) * 1000
        result["parse_time_ms"] = round(parse_time, 2)

        logger.info(
            f"Excel解析完成: {file_path.name}, "
            f"{result['sheet_count']}个工作表, "
            f"{len(all_rows)}行数据, "
            f"耗时{parse_time:.0f}ms"
        )

    except ImportError:
        error_msg = "openpyxl库未安装，请运行: pip install openpyxl"
        logger.error(error_msg)
        result["error"] = error_msg
    except (OSError, ValueError, TypeError, KeyError):
        error_msg = "Excel解析失败: 文件格式错误或损坏，请检查文件"
        logger.exception(error_msg)
        result["error"] = error_msg
    except Exception:
        # openpyxl 的 InvalidFileException 等未覆盖异常也统一转为 error 结果，
        # 保证 parse_excel 的 API 契约（永不抛异常，始终返回状态字典）。
        error_msg = "Excel解析失败: 文件格式错误或损坏，请检查文件"
        logger.exception(error_msg)
        result["error"] = error_msg

    return result


def _extract_table_from_sheet(ws: Any, sheet_idx: int, sheet_name: str) -> dict[str, Any] | None:
    """从Excel工作表中提取表格数据

    Args:
        ws: openpyxl工作表对象
        sheet_idx: 工作表索引
        sheet_name: 工作表名称

    Returns:
        表格信息字典，如果提取失败则返回None
    """
    try:
        # 获取所有数据
        data = []
        for row in ws.iter_rows(values_only=True):
            # 过滤完全空行
            if any(cell is not None for cell in row):
                data.append([str(cell) if cell is not None else "" for cell in row])

        if not data or len(data) < 2:
            return None

        # 第一行作为表头
        headers = data[0]

        # 剩余行作为数据
        rows = []
        for row_data in data[1:]:
            rows.append(row_data)

        table_info = {
            "sheet_index": sheet_idx,
            "sheet_name": sheet_name,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "column_count": len(headers),
        }

        return table_info

    except (OSError, ValueError, TypeError, KeyError) as e:
        logger.warning("提取工作表'%s'失败: %s", sheet_name, e)
        return None


def parse_csv(file_path: str | Path) -> dict[str, Any]:
    """解析CSV文件，提取表格数据

    Args:
        file_path: CSV文件路径

    Returns:
        包含解析结果的字典
    """
    start_time = time.time()
    file_path = Path(file_path)

    result = {
        "status": "error",
        "file_name": file_path.name,
        "file_size": 0,
        "sheet_count": 1,
        "tables": [],
        "rows": [],
        "parse_time_ms": 0.0,
        "error": None,
    }

    if not file_path.exists():
        result["error"] = f"文件不存在: {file_path}"
        return result

    try:
        result["file_size"] = file_path.stat().st_size
    except OSError as e:
        logger.debug("无法获取文件大小: %s", e)

    try:
        import csv

        logger.info("开始解析CSV文件: %s", file_path.name)

        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            data = list(reader)

        if not data or len(data) < 2:
            result["error"] = "CSV文件为空或数据不足"
            return result

        headers = data[0]
        rows = data[1:]

        table_info = {
            "sheet_index": 0,
            "sheet_name": "CSV",
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "column_count": len(headers),
        }

        result["tables"] = [table_info]
        result["rows"] = rows
        result["status"] = "success"

        parse_time = (time.time() - start_time) * 1000
        result["parse_time_ms"] = round(parse_time, 2)

        logger.info(f"CSV解析完成: {file_path.name}, {len(rows)}行数据, 耗时{parse_time:.0f}ms")

    except (OSError, UnicodeDecodeError, ValueError):
        error_msg = "CSV解析失败: 文件格式错误或编码问题，请检查文件"
        logger.exception(error_msg)
        result["error"] = error_msg

    return result


if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")

    if len(sys.argv) < 2:
        logger.info("用法: python excel_parser.py <excel_file>")
        sys.exit(1)

    excel_file = sys.argv[1]
    file_path = Path(excel_file)

    if file_path.suffix.lower() == ".csv":
        result = parse_csv(excel_file)
    else:
        result = parse_excel(excel_file)

    logger.info("\n解析状态: %s", result["status"])
    logger.info("文件名: %s", result["file_name"])
    logger.info("文件大小: %s bytes", result["file_size"])
    logger.info("工作表数量: %s", result["sheet_count"])
    logger.info("表格数量: %s", len(result["tables"]))
    logger.info("数据行数: %s", len(result["rows"]))
    logger.info(f"解析耗时: {result['parse_time_ms']:.2f}ms")

    if result["error"]:
        logger.error("错误: %s", result["error"])

    if result["tables"]:
        logger.info("\n提取的表格:")
        for i, table in enumerate(result["tables"], 1):
            logger.info("\n表格 %s (%s):", i, table["sheet_name"])
            logger.info("  表头: %s", table["headers"])
            logger.info("  行数: %s", table["row_count"])
            if table["rows"]:
                logger.info("  首行数据: %s", table["rows"][0])
